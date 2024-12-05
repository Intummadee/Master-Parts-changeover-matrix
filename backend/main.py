from fastapi import FastAPI, HTTPException, Depends, File, UploadFile, Response
import shutil
from sqlmodel import Session, select
import json
import pathlib
from typing import List, Union  
from database import engine, Parts , Changeover 
from fastapi.responses import FileResponse , StreamingResponse, JSONResponse
from io import BytesIO
import pandas as pd

from openpyxl.styles import PatternFill

import random

# BytesIO ใช้สำหรับจัดการข้อมูลในหน่วยความจำแทนการสร้างไฟล์จริง
# FastAPI และ Depends: ใช้สำหรับสร้าง API และจัดการ dependency injection
# StreamingResponse: ใช้ส่งข้อมูลที่สร้างขึ้น (เช่น ไฟล์ Excel) กลับไปยังผู้ใช้
# pandas: ใช้สำหรับสร้าง DataFrame และจัดการข้อมูลในรูปแบบตาราง
# Session: ใช้จัดการการเชื่อมต่อฐานข้อมูล SQLAlchemy
# PatternFill: มาจาก openpyxl ใช้สำหรับเติมสีพื้นหลังในเซลล์ Excel

def random_color():
    return ''.join(random.choices('0123456789ABCDEF', k=6))
    # ชุดตัวอักษรนี้ประกอบด้วยตัวเลข (0–9) และตัวอักษร A–F ซึ่งเป็นชุดตัวอักษรที่ใช้ในรหัสสีแบบ HEX
    # k=6: ระบุจำนวนครั้งที่ต้องสุ่ม คือ 6 ครั้ง (6 ตัวอักษร) ซึ่ง รหัสสี HEX จะมีความยาว 6 ตัวอักษร เช่น #FF5733

import datetime

# Excel
import pandas as pd
from io import BytesIO


app = FastAPI()

from fastapi.middleware.cors import CORSMiddleware
# Allow CORS for your frontend
app.add_middleware(
    CORSMiddleware,
    # allow_origins=["http://localhost:8080"],  # frontend origin
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],  # Allow all HTTP methods (GET, POST, etc.)
    allow_headers=["*"],  # Allow all headers
)

data = []

# define app start-up event
@app.on_event("startup")
async def startup_event():
    DATAFILE = pathlib.Path() / 'data' / 'allParts.json'

    session = Session(engine)
    stmt = select(Parts)
    result = session.exec(stmt).first()

    # load data if there's no result
    if result is None:
        with open(DATAFILE, 'r') as f:
            allParts = json.load(f)
            for eachPart in allParts:
                session.add(Parts(**eachPart))
        session.commit()


    # Load changeover data
    DATAFILE_CHANGEOVER = pathlib.Path() / 'data' / 'changeOver.json'

    # Check if Changeover table is empty, then load changeover data
    stmt_changeover = select(Changeover)
    result_changeover = session.exec(stmt_changeover).first()
    if result_changeover is None:
        with open(DATAFILE_CHANGEOVER, 'r') as f:
            changeovers = json.load(f)
            for changeover in changeovers:
                session.add(Changeover(**changeover))
        session.commit()


    session.close()


def get_session():
    with Session(engine) as session:
        yield session



@app.get("/parts", response_model=List[Parts])  # Changed Track to Parts here
def parts(session: Session = Depends(get_session)): 
    stmt = select(Parts)
    result = session.exec(stmt).all()
    return result



@app.get("/parts/{parts_id}", response_model=Union[Parts, str])
def parts(parts_id: int , response: Response, session: Session = Depends(get_session)): 
    part = session.get(Parts , parts_id)
    if parts is None:
        response.status_code = 404
        return "Not found this Part"
    return part


# Post
@app.post("/parts/", response_model=Parts, status_code=201) 
def create_parts(part: Parts, session: Session = Depends(get_session)):
    # ตรวจสอบว่า part_name ไม่เป็นค่าว่าง
    if not part.part_name or part.part_name.strip() == "":
        raise HTTPException(status_code=400, detail="Part name is required")

    # ตรวจสอบว่า part_name ไม่ซ้ำ
    existing_part = session.query(Parts).filter(Parts.part_name == part.part_name).first()
    if existing_part:
        raise HTTPException(status_code=400, detail="Part name already exists")

    # สร้างข้อมูล part ใหม่
    new_part = Parts(part_name=part.part_name)

    # เพิ่มข้อมูลลงในฐานข้อมูล
    session.add(new_part)
    session.commit()
    session.refresh(new_part)

    return new_part

# Put
@app.put("/parts/{parts_id}", response_model=Union[Parts, str])  # Changed Track to Parts here
def update_parts(parts_id: int , updated_part: Parts , response: Response, session: Session = Depends(get_session)): 
    part = session.get(Parts , parts_id)

    if part is None:
        response.status_code = 404
        return "Parts not found"
    
    # update
    part_dict = updated_part.dict(exclude_unset=True)
    for key , val in part_dict.items():
        setattr(part, key , val)
    session.add(part)
    session.commit()
    session.refresh(part)

    return part


# Delete
@app.delete("/parts/{parts_id}")  # Changed Track to Parts here
def delete_parts(parts_id: int ,response: Response, session: Session = Depends(get_session)): 
    
    part = session.get(Parts , parts_id)

    if part is None:
        response.status_code = 404
        return "Parts not found"
    
    session.delete(part)
    session.commit()
    return Response(status_code=200)




@app.get("/changeOver", response_model=List[Changeover])  # Changed Track to Parts here
def parts(session: Session = Depends(get_session)): 
    stmt = select(Changeover)
    result = session.exec(stmt).all()
    return result



# Delete
@app.delete("/changeOver/{from_part_id}/{to_part_id}")
def delete_changeover(from_part_id: int, to_part_id: int, response: Response, session: Session = Depends(get_session)):
    stmt = select(Changeover).where(Changeover.from_part_id == from_part_id, Changeover.to_part_id == to_part_id)
    changeover = session.exec(stmt).first()
    
    if not changeover:
        response.status_code = 404
        return "Changeover entry not found"
    
    session.delete(changeover)
    session.commit()
    return Response(status_code=200)



# Post
@app.post("/changeOver/", response_model=Changeover, status_code=201)  # Changed Track to Parts here
def create_parts(Changeover: Changeover, session: Session = Depends(get_session)): 
    session.add(Changeover)
    session.commit()
    session.refresh(Changeover)
    return Changeover







# ปุ่ม upload แล้วแก้ฐานข้อมูลให้ตรงกัน
@app.post("/export-to-excel")
async def export_to_excel(data: list):
    df = pd.DataFrame(data)

    excel_file = BytesIO()
    df.to_excel(excel_file, index=False)
    excel_file.seek(0)

    

    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=parts_changeover.xlsx"
        }
    )



# ปุ่มสร้าง excel จากฐานข้อมูล 
@app.get("/table", response_model=List)
# รับพารามิเตอร์ session เพื่อเข้าถึงฐานข้อมูลโดยใช้ SQLAlchemy
def get_table(session: Session = Depends(get_session)):
    parts = session.query(Parts).all()
    changeovers = session.query(Changeover).all()

    table_data = []

    for part in parts:
        changeover_times = {} # จะมีการสร้าง dictionary สำหรับ changeover_times

        for changeover in changeovers:
            if changeover.from_part_id == part.id: # ใช้ from_part_id และ to_part_id เพื่อจับคู่ชิ้นส่วน
                to_part = next((p for p in parts if p.id == changeover.to_part_id), None) 
                # เป็นการค้นหา ชิ้นส่วน (part) ในลิสต์ parts ที่มี id ตรงกับ to_part_id ของ changeover
                if to_part:
                    changeover_times[to_part.part_name] = changeover.changeover_time

        table_data.append({
            "id": part.id,
            "part_name": part.part_name,
            "changeoverTimes": changeover_times
        })

    
    df = pd.DataFrame([
        {
            "Part Name": part["part_name"],
            **part["changeoverTimes"]
            # ใช้สำหรับการ unpacking หรือ ขยาย dictionary ที่อยู่ใน part["changeoverTimes"] ให้ออกมาเป็นคู่ของ key และ value ซึ่งจะนำไปใช้ใน dictionary ใหม่ที่กำลังสร้าง
        }
        for part in table_data
    ])

    # คอลัมม
    # ในขั้นตอนนี้ คอลัมน์ที่ชื่อขึ้นต้นด้วย TG จะถูกจัดเรียงตามลำดับตัวเลขที่ตามหลัง TG (เช่น TG1, TG2, TG3 ฯลฯ) 
    tg_columns = sorted([col for col in df.columns if col.startswith("TG")], key=lambda x: int(x[2:]))
    # key: ใช้เพื่อระบุว่าค่าหลักใดจะนำมาใช้เป็นเกณฑ์ในการเรียงลำดับ
    # lambda x: int(x[2:]): เป็นฟังก์ชันแบบย่อ (anonymous function) ที่รับค่าหนึ่ง (x) และคืนค่าที่ต้องการใช้ในการเปรียบเทียบ
    # x[2:]: ตัดอักขระสองตัวแรกของสตริง x ออก (เริ่มนับจาก index 2)
    # int(): แปลงส่วนที่เหลือให้เป็นตัวเลข (ชนิด int)
    ordered_columns = ["Part Name"] + tg_columns # จัดลำดับใหม่ โดยให้ "Part Name" อยู่คอลัมน์แรก ตามด้วยคอลัมน์ "TG..."
    df = df.loc[:, ordered_columns] # ปรับ DataFrame ให้ใช้ลำดับคอลัมน์ตามที่กำหนด

    print("--------------------------------------------")
    print(df)

    # ใช้ Pandas ExcelWriter และ openpyxl engine เพื่อสร้างไฟล์ Excel จาก DataFrame
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
    # ExcelWriter: ใช้เขียน DataFrame ลงในไฟล์ Excel
        df.to_excel(writer, index=False, sheet_name='Parts Table')

        # โหลด workbook เพื่อแก้ไขไฟล์ Excel
        workbook = writer.book
        worksheet = writer.sheets['Parts Table'] # writer.sheets['Parts Table'] ดึง worksheet ชื่อ "Parts Table"

        column_colors = {}
        


        # PatternFill กำหนดสีพื้นหลัง (ในตัวอย่างคือสีเหลือง FFFF00)
        for col in tg_columns:

            # สุ่มสีสำหรับแต่ละคอลัมน์
            random_hex = random_color()
            highlight_fill = PatternFill(start_color=random_hex, end_color=random_hex, fill_type="solid")
            # print("highlight_fill ,  ", highlight_fill ) highlight_fill ,   <openpyxl.styles.fills.PatternFill object>
            # เก็บสีในพจนานุกรม
            column_colors[col] = random_hex

            col_idx = df.columns.get_loc(col) + 1 # ต้องเพิ่ม 1 เพราะ Excel ใช้ index เริ่มต้นที่ 1
            # for row_idx in range(2, len(df) + 2): # เริ่มที่ 2 เพื่อข้าม header
            worksheet.cell(row=1, column=col_idx).fill = highlight_fill

            for row_idx in range(2, len(df) + 2):  # เริ่มที่แถว 2
                cell_value = worksheet.cell(row=row_idx, column=1).value  # ดึงค่าจากคอลัมน์ "Part Name"
                if cell_value == col:  # ถ้าชื่อแถว (ในคอลัมน์ "Part Name") ตรงกับชื่อคอลัมน์
                    worksheet.cell(row=row_idx, column=1).fill = highlight_fill
                    worksheet.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="BEBEBE", end_color="BEBEBE", fill_type="solid")

        
         


    output.seek(0) # ตั้ง pointer ไปที่จุดเริ่มต้นของไฟล์ในหน่วยความจำ

    

    return StreamingResponse(
        # StreamingResponse: ใช้สำหรับส่งข้อมูลจากหน่วยความจำ เช่น BytesIO แทนการใช้ไฟล์ที่อยู่ในระบบ

        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=parts_table.xlsx'}
        # Content-Disposition เป็น header ที่ระบุว่าไฟล์ที่ส่งกลับควรถูกดาวน์โหลดเป็นไฟล์ชื่อ parts_table.xlsx
        
    )



#         #! excel_data
#         # excel_data    Part Name   TG1     TG2     TG3
#         # 0             TG1         Null    5.5     2.0
#         # 1             TG2         3.5     Null    8.0
#         # 2             TG3         22.0    12.0    Null
#         # 3             ss          1.0     2.0     3.0

            # [{TG1 , "#ddd"}]

#         #! data
#         # data [
#         # {'Part Name': 'TG1', 'TG1': 'Null', 'TG2': 5.5, 'TG3': 2.0}, 
#         # {'Part Name': 'TG2', 'TG1': 3.5, 'TG2': 'Null', 'TG3': 8.0}, 
#         # {'Part Name': 'TG3', 'TG1': 22.0, 'TG2': 12.0, 'TG3': 'Null'}, 
#         # {'Part Name': 'ss', 'TG1': 1.0, 'TG2': 2.0, 'TG3': 3.0}] # เพิ่มบรรทัดนี้ลง Excel

# Upload Excel from front and update dB
@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...), session: Session = Depends(get_session)):
    try:
        # Read the Excel file
        contents = await file.read()
        excel_data = pd.read_excel(BytesIO(contents), engine='openpyxl')
        
        # Replace NaN with None to make the data JSON compliant
        excel_data = excel_data.where(pd.notnull(excel_data), "Null")
        
        print("excel_data ", excel_data)

        # Convert the data to a list of dictionaries
        data = excel_data.to_dict(orient="records")
        with Session(engine) as session:
            # Step 1: Sync parts from Excel to DB
            part_names_in_excel = {entry['Part Name'] for entry in data}
            
            # Fetch all existing parts from DB
            existing_parts = session.exec(select(Parts)).all()
            existing_part_names = {part.part_name for part in existing_parts}
            
            # Add new parts
            new_parts = part_names_in_excel - existing_part_names
            for new_part_name in new_parts:
                session.add(Parts(part_name=new_part_name))
            
            # Remove parts not in Excel
            parts_to_remove = existing_part_names - part_names_in_excel
            for part in existing_parts:
                if part.part_name in parts_to_remove:
                    session.delete(part)
            
            # Step 2: Update Changeover Times
            for entry in data:
                from_part_name = entry['Part Name']
                from_part = session.exec(select(Parts).where(Parts.part_name == from_part_name)).first()
                
                for key, value in entry.items():
                    if key != 'Part Name':
                        to_part_name = key
                        to_part = session.exec(select(Parts).where(Parts.part_name == to_part_name)).first()
                        
                        if to_part:
                            changeover_time = 0.0 if value == 'Null' else value
                            
                            # Check if changeover record exists
                            changeover = session.exec(
                                select(Changeover).where(
                                    (Changeover.from_part_id == from_part.id) &
                                    (Changeover.to_part_id == to_part.id)
                                )
                            ).first()
                            
                            if changeover:
                                # Update existing record
                                if changeover.changeover_time != changeover_time:
                                    changeover.changeover_time = changeover_time
                            else:
                                # Create new changeover record
                                session.add(Changeover(
                                    from_part_id=from_part.id,
                                    to_part_id=to_part.id,
                                    changeover_time=changeover_time
                                ))
            session.commit()
            return {"status": "success", "message": "Data processed and added/updated in the database"}
    except Exception as e:
        print("Error:", str(e))
        return JSONResponse(content={"status": "error", "message": str(e)}, status_code=500)